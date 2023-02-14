"""
Detailed documentation and code examples coming soon.
"""
from __future__ import annotations
# External
import re
import pandas as pd
from typing import Any, overload
from copy import deepcopy
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color

# Internal main
from excelbird._layout_references import Globals
from excelbird._base.identifier import HasId
from excelbird._base.styling import HasBorder
from excelbird._base.dotdict import Style
from excelbird._base.loc import Loc

from excelbird.formats import number_formats as num_formats

from excelbird._utils.util import (
    get_dimensions,
)
from excelbird._utils.cell_util import (
    autofit_algorithm,
    remove_paren_enclosure,
    prefix_formulae_funcs,
    format_formula,
)
from excelbird._utils.color_algorithms import (
    color_is_light,
    get_alt_shade,
)
from excelbird.exceptions import AlreadyWrittenError, CellReferenceError
from excelbird._base.math import CanDoMath
from excelbird.core.expression import Expr
from excelbird.core.function import Func

cell_reference_warning_issued = False


class Cell(HasId, HasBorder, CanDoMath):
    """
    Represents a real cell in Excel. Capable of math operations in python, and tracking
    cell references on its own, without knowing locations.

    Excelbird cells are *not* mere containers for data. Since they represent a real object in Excel,
    each can only exist *once* inside a Book. This is because when you reference a Cell in a python
    expression (i.e. `cell_c = cell_a + cell_b`), cell references will be established so that when
    `cell_c` is written to the workbook, its value will be a formula that references the real
    locations of `cell_a` and `cell_b`. So if `cell_a` is found in multiple places in a Book,
    this formula could not be resolved.

    ALL attributes have a default of None. This allows parent containers to know whether a value
    has been set or not, to avoid override. For any attribute, if you just want to ignore the
    value set by a parent container, set its value to False, even if it isn't documented as accepting
    a boolean.

    Parameters
    ----------

    value : str or int or float, optional
        Value to display. If None, the Cell's styling will NOT be rendered. To display an empty
        cell with styling, pass an empty string as a value.
        This attribute will temporarily be None if a Cell references other Cells in an expression.
        For instance, if `cell3 = cell1 + cell2`, cell3's value will be None, until the parent
        book's .write() is called and cell1 and cell2's locations have been determined.
    dropdown : list or Cell or Col or Row, optional
        Apply data validation for the cell that offers a dropdown list of values to pick from. Pass
        a list of options, or a Cell, Col, or Row.
    id : str, optional
        Unique identifier to store globally so that this element can be referenced
        elsewhere in the layout without being assigned to a variable
    align_x : str, optional
        Horizontal alignment. Options: 'center', 'right', 'left'
    align_y : str, optional
        Vertical alignment. Options: 'top', 'center', 'bottom'
    indent : int or float, optional
        Indent the cell's value. Indentation direction will depend on horizontal alignment, so if
        align_x='right', the indentation will determine its distance from the right edge of the cell.
    center : bool, optional
        Center the cell horizontally and vertically. Shorthand for setting align_x and align_y to 'center'
    wrap : bool, optional
        Wrap the cell's text so that it doesn't continue outside of its boundary. As long as row_height=None
        and col_width=None, Excel will automatically resize the cell to display the full text.
    size : int, optional
        Font size
    bold : bool, optional
        Bold font
    italic : bool, optional
        Italic font
    color : str, optional
        Font color, as a hex code
    num_fmt : str, optional
        Cell value format. See the styles module for available formats
    currency : bool, optional
        Indicate that this Cell might contain a currency value. If value is an int or float
        and num_fmt has not been set, apply accounting format. The recommended use-case for this
        attribute is to set currency=True for a parent container once, and all of its numerical data
        will be formatted accordingly
    ignore_format : bool, optional
        Negate any number formatting set by a parent container. This is an easier alternative to
        setting num_fmt=False, currency=False
    fill_color : str, optional
        Hex code string color to fill the cell
    auto_color_font : bool, optional
        Sets font color to white or black, based on lightness of fill_color, so that text will be
        visible over any background. Lightness of fill_color is measured using the weighted euclidean
        norm of the rgb vector. If the resulting coefficient indicates a medium to light color, lightness
        will be re-calculated from the Luma of the rgb vector.
    auto_shade_font : bool, optional
        Font color will be a lighter or darker shade of fill_color. If fill_color is dark, a lighter shade
        will be chosen as the font color, and vice versa. Lightness coefficient of fill_color is measured
        using the weighted euclidean norm of the rgb vector. If the resulting coefficient indicates a
        medium to light color, lightness will be re-calculated from the Luma of the rgb vector.
    border : list[tuple or str or bool] or tuple[str or bool, str or bool] or str or bool, optional
        Syntax inspired by CSS. A non-list value will be applied to all 4 sides. If list,
        length can be 2, 3, or 4 elements. Order is [top, right, bottom, left]. If length 2,
        apply the first element to top and bottom border, and apply the second element to right and left.
    border_top : tuple[str or bool, str or bool] or str or bool, optional
        Top border. If True, a thin black border is used. If string (6 char hex code),
        use the default weight and apply the specified color. If string (valid weight name),
        use the default color and apply the specified weight. If tuple, apply the first
        element as weight, and second element as color.
    border_right : tuple[str or bool, str or bool] or str or bool, optional
        Right border. See border_top
    border_bottom : tuple[str or bool, str or bool] or str or bool, optional
        Bottom border. See border_top
    border_left : tuple[str or bool, str or bool] or str or bool, optional
        Left border. See border_top
    col_width : int, optional
        Column width. Format is the same as used in Excel.
    row_height : int, optional
        Row height. Format is the same as used in Excel.
    autofit : bool, optional
        Autofit column width based on the length of the value. This is NOT as accurate as the built-in
        autofit feature in Excel. This is because we can't determine the rendered width without actually
        rendering the value in the desired font and size, and counting pixels.
    merge : tuple[int, int], optional
        Merge this cell with other cells to its right or below. First element is the distance to
        merge below, and second element is the distance to merge across. For instance, `(0, 1)` will
        merge the current Cell with the one to the right. `(1,1)` will merge diagonally in a square.
        The cells being merged must have values of None. For instance, if you have a Row of multiple
        values and want the first and second elements to be merged, your code would be as follows:
        ``Row(Cell('a', merge=(0,1)), Cell(), Cell('b'), Cell('c'))`` - notice the second Cell is empty
    cell_style : dict, optional
        A dict that contains attributes to set. Priority is given to existing attributes - An attribute in
        cell_style will only be set if the Cell's attribute is currently None

    """

    _dimensions = 0
    elem_type = None

    @overload
    def __new__(cls, fn: str | Func, **kwargs) -> Func:
        ...

    @overload
    def __new__(cls, func: str | Func, **kwargs) -> Func:
        ...

    @overload
    def __new__(cls, ex: str | set | Expr, **kwargs) -> Expr:
        ...

    @overload
    def __new__(cls, expr: str | set | Expr, **kwargs) -> Expr:
        ...

    @overload
    def __new__(cls, *args, **kwargs) -> Cell:
        ...

    def __new__(cls, *args, fn=None, func=None, ex=None, expr=None, **kwargs):
        fn = fn if fn is not None else func
        ex = ex if ex is not None else expr
        if isinstance(fn, Func):
            fn = fn.inner
        if isinstance(ex, Expr):
            ex = ex.expr_str

        if fn is not None:
            new_func = Func.__new__(Func)
            new_func.__init__(fn, res_type=cls, **kwargs)
            return new_func

        if ex is not None:
            new_expr = Expr.__new__(Expr)
            new_expr.__init__(ex, res_type=cls, **kwargs)
            return new_expr

        return super().__new__(cls)

    def __init__(
        self,
        value: Any | None = None,
        dropdown: list | Any | None = None,  # list of values, Col, Row, or Cell
        id: str | None = None,
        align_x: str | None = None,
        align_y: str | None = None,
        indent: float | None = None,
        center: bool | None = None,  # center align horizontal and vertical
        wrap: bool | None = None,
        size: int | None = None,  # font size
        bold: bool | None = None,
        italic: bool | None = None,
        color: str | None = None,
        num_fmt: str | None = None,
        currency: bool | None = None,
        ignore_format: bool | None = None,
        fill_color: str | None = None,
        auto_color_font: bool | str | None = None,
        auto_shade_font: bool | float | None = None,
        border_left: bool | str | None = None,
        border_right: bool | str | None = None,
        border_top: bool | str | None = None,
        border_bottom: bool | str | None = None,
        border: bool | str | None = None,  # MUST be last border attr set
        col_width: int | None = None,  # finds the cell's column and adjust it
        row_height: int | None = None,  # finds the cell's row and adjusts it
        merge: tuple[int, int] | None = None,  # (y, x) counts of how many merges. Ex: (0, 1) merges cell with right adjacent
        _expr: list | None = None,
        _func: list | None = None,
        autofit: bool | None = None,
        cell_style: dict | None = None,
        _written: bool | None = None,
        fn: str | Func | None = None,
        func: str | Func | None = None,
        ex: str | set | Expr | None = None,
        expr: str | set | Expr | None = None,
    ) -> None:
        del fn
        del func
        del ex
        del expr
        self._written = False
        self._loc = None

        self.value = value
        self.dropdown = dropdown
        self.id = id
        self.align_x = align_x
        self.align_y = align_y
        self.indent = indent
        self.wrap = wrap
        self.size = size
        self.bold = bold
        self.italic = italic
        self.color = color
        self.num_fmt = num_fmt
        self.currency = currency
        self.ignore_format = ignore_format
        self.fill_color = fill_color
        self.auto_color_font = auto_color_font
        self.auto_shade_font = auto_shade_font
        self.col_width = col_width
        self.row_height = row_height
        self.merge = merge
        self._expr = _expr
        self._func = _func
        self.autofit = autofit
        self.center = center

        self._init_border(
            border,
            border_top,
            border_right,
            border_bottom,
            border_left,
        )
        self._inherit_style_without_override(cell_style)

        if isinstance(self.value, Cell):
            cell = self.value
            self.value = None
            new_dict = {
                k: v for k, v in cell.__dict__.items() if k not in ["_id", "_loc"]
            }
            for key, val in new_dict.items():
                if getattr(self, key) is None:
                    setattr(self, key, val)

        elif isinstance(self.value, (Expr, Func)):
            raise ValueError(
                "Can't pass Expr or Func as a Cell's `value`. "
                "Instead, use the expression/function by itself, and pass any "
                "Cell styling as a dict to `cell_style`."
            )

    @property
    def is_empty(self) -> bool:
        return self.value is None and self._func is None and self._expr is None

    @property
    def shape(self) -> tuple:
        return tuple()

    @property
    def width(self) -> int:
        return 1

    @property
    def height(self) -> int:
        return 1

    def ref(self, inherit_style: bool = False, **kwargs) -> Cell:
        """
        Get a new `Cell` who will display a cell reference to the current
        cell. This assumes that **both** the calling object
        and the returned object will be placed in the workbook.

        .. note::

            Calling ``.ref()`` is **not** necessary when an object is used in
            a python expression (i.e. ``some_cell + some_row``) and should `only`
            be used to duplicate data across a workbook.

        Parameters
        ----------
        inherit_style : bool, default False
            Copy the caller's style to the returned object.
        **kwargs : Any
            Extra keyword arguments are set as attributes on the returned
            object.

        Returns
        -------
        :class:`Cell <excelbird.Cell>`

        """
        if inherit_style is True:
            self_dict = deepcopy(self.__dict__)
            for key, val in self_dict.items():
                if key not in kwargs and key not in ["_id", "_loc", "_expr", "_func"]:
                    kwargs[key] = val
        return Cell(_expr=[self], **kwargs)

    def _expr_value(self) -> str | None:
        if self._expr is None:
            return None
        expr = str(self._eval_expr(self._expr))
        return remove_paren_enclosure(expr)

    def _func_value(self) -> str | None:
        if self._func is None:
            return None
        return self._eval_func(self._func)

    def _write(self) -> None:

        assert self._loc is not None, (
            "Excelbird developer error: Somehow, ._write() got called "
            "on a Cell that doesn't have a location. This is a serious issue!"
        )

        if self._written is True:
            raise AlreadyWrittenError(
                "Excelbird objects can only be written to a workbook once. This is "
                "because when `.write()` is called on a `Book`, the state of each of its elements "
                "changes: expressions are evaluated, series headers are inserted as actual cells, "
                "and cells with references are filled with the string locations of their references. "
                "\nSupport for repeated writes is possible, but hasn't been implemented yet."
            )

        if self._func is not None:
            self.value = "=" + self._func_value()
            self.value = self.value.replace(self._loc.title_str, "")
            self.value = prefix_formulae_funcs(self.value)
            self.value = format_formula(self.value)

        if self._expr is not None:
            self.value = self._expr_value()
            if self.value is None:
                return
            if "UNKNOWN" not in str(self.value):
                self.value = "=" + str(self.value)
            self.value = self.value.replace(self._loc.title_str, "")

        if self.value is None:
            return

        try:
            if pd.isnull(self.value):
                return
        except Exception:
            # Just making sure pd.isnull doesn't throw, if given a data type it doesn't like
            pass

        y, x = self._loc.y, self._loc.x
        cell = self._loc.cell
        cell.value = self.value

        # if ":" in str(self.value):
        #     self._loc.ws.formula_attributes['A5'] = {'t': 'array', 'ref': "A5:A5"}

        def get_dropdown() -> DataValidation | None:
            value = self.dropdown
            if value is None:
                return

            if type(value) in [list, tuple]:
                dropdown_items = [str(x) for x in value]
                formula = '"' + ",".join(dropdown_items) + '"'

            else:
                from excelbird.core.series import Col, Row

                if not isinstance(value, (Cell, Col, Row)):
                    raise ValueError(f"Invalid type for dropdown, {type(value)}")

                formula = None
                if isinstance(value, (Col, Row)):
                    formula = self._eval_expr(value.range()._expr).replace(
                        self._loc.title_str, ""
                    )
                else:
                    if value._loc is None:
                        raise ValueError(
                            "Cell reference in dropdown must be a valid Cell in workbook"
                        )

                    formula = self._eval_expr([value]).replace(self._loc.title_str, "")

            if formula is None:
                return None

            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(cell)
            return dv

        validation = get_dropdown()
        if validation is not None:
            self._loc.ws.add_data_validation(validation)

        def get_number_format():
            if self.ignore_format is True:
                return

            if isinstance(self.num_fmt, str):
                return self.num_fmt

            if isinstance(self.value, str):
                return

            if isinstance(self.value, float):
                if self.currency is True:
                    return num_formats.accounting_float
                return num_formats.comma_float

            if isinstance(self.value, int):
                if self.currency is True:
                    return num_formats.accounting_int
                return num_formats.comma_int

        number_format = get_number_format()
        if number_format is not None:
            cell.number_format = number_format

        align, font, fill, border = {}, {}, {}, {}

        if self.center is True:
            align["horizontal"] = "center"
            align["vertical"] = "center"

        if self.align_x is not None:
            align["horizontal"] = self.align_x
        if self.align_y is not None:
            align["vertical"] = self.align_y

        if isinstance(self.indent, (int, float)):
            align["indent"] = self.indent

        if self.wrap is not None:
            align["wrap_text"] = self.wrap

        if self.size is not None:
            font["size"] = self.size
        if self.bold is not None:
            font["bold"] = self.bold
        if self.italic is not None:
            font["italic"] = self.italic
        if self.color is not None:
            font["color"] = self.color
        else:
            if self.auto_color_font is True and isinstance(self.fill_color, str):
                if not color_is_light(self.fill_color):
                    font["color"] = "FFFFFF"  # white
            elif self.auto_shade_font is not None and isinstance(self.fill_color, str):
                if isinstance(self.auto_shade_font, float):
                    font["color"] = get_alt_shade(self.fill_color, self.auto_shade_font)
                else:
                    font["color"] = get_alt_shade(self.fill_color)

        if self.fill_color is not None:
            fill = {"patternType": "solid", "fgColor": Color(self.fill_color)}

        def get_border(border) -> dict:
            def get_side(side) -> Side:
                if (
                    side is None
                    or side is False
                    or side == (None, None)
                    or side == (False, False)
                ):
                    return None
                assert isinstance(
                    side, tuple
                ), f"Internal developer error processing border. Border side value, {side} is invalid"
                if isinstance(side[0], str) and isinstance(side[1], str):
                    return Side(style=side[0], color=side[1])
                if isinstance(side[0], str):
                    return Side(style=side[0])
                if isinstance(side[1], str):
                    return Side(style=HasBorder.default_weight, color=side[1])
                raise ValueError(side)

            res = {}
            top, right, bottom, left = border
            if (side_top := get_side(top)) is not None:
                res["top"] = side_top
            if (side_right := get_side(right)) is not None:
                res["right"] = side_right
            if (side_bottom := get_side(bottom)) is not None:
                res["bottom"] = side_bottom
            if (side_left := get_side(left)) is not None:
                res["left"] = side_left
            return res

        border = get_border(self.border)

        if len(font) > 0:
            cell.font = Font(**font)
        if len(fill) > 0:
            cell.fill = PatternFill(**fill)
        if len(border) > 0:
            cell.border = Border(**border)
        if len(align) > 0:
            cell.alignment = Alignment(**align)

        if self.merge is not None:
            end_row = 1 + y + self.merge[0]
            end_column = 1 + x + self.merge[1]
            self._loc.ws.merge_cells(
                start_row=y + 1,
                start_column=x + 1,
                end_row=end_row,
                end_column=end_column,
            )

        if self.col_width is not None:
            self._loc.column_dimensions.width = self.col_width

        if self.autofit is True and self.col_width is None:
            curr = self._loc.column_dimensions.width
            new = autofit_algorithm(self.value)
            if new > curr:
                self._loc.column_dimensions.width = new

        if self.row_height is not None:
            self._loc.row_dimensions.height = self.row_height

        self._written = True

    def _set_loc(self, loc: Loc) -> None:
        self._loc = loc

    def __repr__(self):
        if self._expr is not None:
            return type(self).__name__ + "({...})"
        if self._func is not None:
            return type(self).__name__ + "(Func(...))"
        return f"{type(self).__name__}({self.value})"

    def _eval_func(self, func: list) -> str:
        def format_element(elem) -> str:
            if isinstance(elem, str):
                return elem

            if isinstance(elem, (int, float)):
                return str(elem)

            if get_dimensions(elem) > 0:
                cell_range = elem.range()._expr
                evaluated = self._eval_expr(cell_range)
                return remove_paren_enclosure(evaluated)

            if elem._loc is not None:
                return elem._loc.full_str

            if elem._expr is not None:
                evaluated = self._eval_expr(elem._expr)
                return remove_paren_enclosure(evaluated)

            if elem.value is not None:
                return str(elem.value)  # Don't put quotes around strings here

        res = "".join([format_element(e) for e in func])
        return res

    def _eval_expr(self, expr: list) -> str:
        def format_element(elem) -> str:
            if not isinstance(elem, Cell):
                return str(elem)

            if elem._loc is not None:
                return elem._loc.full_str

            if elem._expr is not None:
                return self._eval_expr(elem._expr)

            if elem._func is not None:
                return self._eval_func(elem._func)
            else:
                global cell_reference_warning_issued

                if elem.value is not None:
                    # if cell_reference_warning_issued is False:
                    #     print(
                    #         "Warning: A cell in your book is trying to reference a cell which "
                    #         "won't be placed in the book. The missing cell's value has been applied "
                    #         "as a hardcoded value in the valid cell's expression."
                    #     )
                    #     cell_reference_warning_issued = True

                    if isinstance(elem.value, str):
                        quote_stripped_val = elem.value.strip('"')
                        return f'"{quote_stripped_val}"'
                    else:
                        return str(elem.value)

                if (
                    cell_reference_warning_issued is False
                    and Globals.force_valid_references is False
                ):
                    CellReferenceError.issue_warning()
                    cell_reference_warning_issued = True
                else:
                    raise CellReferenceError()

                return "UNKNOWN"

        res = [format_element(e) for e in expr]

        if len(res) > 2:
            res = ["("] + res + [")"]

        return "".join(res)

    def _inherit_style_without_override(self, new_style: dict | Style | None) -> None:
        if new_style is not None:
            for key, val in new_style.items():
                check_unset = lambda x: x is None
                if key == "border":
                    check_unset = lambda x: x == [None, None, None, None]

                if check_unset(getattr(self, key, None)):
                    setattr(self, key, val)
