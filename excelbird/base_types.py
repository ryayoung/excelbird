from typing import Any, overload, TypeVar, Iterable
from types import NoneType

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from excelbird.globals import Globals


class ImpliedType:
    """
    Short for 'Implied' type. Serves as a temporary container to store arguments
    to later resolve to a desired type

    Once created, an `I` cannot:
        be modified
        be used in expressions/functions
        have its elements/attributes referenced
    until it is either:
        passed as an element to a parent container
        resolved by passing a type to its `resolve_to()` function

    Main purpose (other than for conciseness) is to let you swap parent container types
    without having to refactor element types. For instance, if your code looks like:
        HFrame(Col(...), Col(...))
    and you want to change HFrame to VFrame, you would have to track down every Col element
    and refactor it to a Row. Instead:
        HFrame(I(...), I(...)) easily refactors to: VFrame(I(...), I(...))
    """

    def __init__(self, *args, **kwargs) -> None:
        self.__args = args
        self.__kwargs = kwargs

    def astype(self, dtype: type, **kwargs) -> Any:
        return dtype(*self.__args, **self.__kwargs, **kwargs)

    @classmethod
    def resolve_all_in_container(cls, container: list, dtype: type):
        for i, elem in enumerate(container):
            if isinstance(elem, cls):
                container[i] = elem.astype(dtype)


class I(ImpliedType):
    """
    Shorthand for `ImpliedType`
    """

    pass


class Gap(int):
    def __new__(cls, value=None, *args, **kwargs):
        if value is None:
            value = 1
        # Need this because Gap's __init__ takes extra args, which int's
        # __new__ doesn't accept. So just call int's __new__ with the `value`
        # arg and ignore the extras which will be handled by Gap __init__
        return super(Gap, cls).__new__(cls, value)

    def __init__(self, value: int | None = None, fill: bool = False, is_margin: bool = False, **kwargs):
        if value is None:
            value = 1
        if len(kwargs) > 0:
            fill = True
        self.fill = fill
        self.is_margin = is_margin
        self.kwargs = kwargs
        int.__init__(value)

    def __len__(self):
        return self

    @property
    def fill_val(self):
        if self.fill is True:
            return ""
        return None

    def ref(self):
        return self

    @property
    def width(self) -> int:
        return int(self)

    @property
    def height(self) -> int:
        return int(self)

    @classmethod
    def explode_all_to_values(cls, container: list, val_type: type) -> None:
        """
        Given a container, explode each Gap to `val_type`
        with the gap's fill_val

        Mutates inplace: `container`
        """
        for i, elem in enumerate(container):
            if isinstance(elem, cls):
                gap = container.pop(i)
                for _ in range(gap):
                    container.insert(i, val_type(gap.fill_val, **gap.kwargs))

    @classmethod
    def explode_all_to_vecs(
        cls, container: list, vec_type: type, vec_length: int
    ) -> None:
        """
        Given a container, explode each Gap to vectors of vec_type filled with
        `val_type` with the Gap's fill_val

        Mutates inplace: `container`
        """
        val_type = vec_type.elem_type
        for i, elem in enumerate(container):
            if isinstance(elem, cls):
                gap = container.pop(i)
                for _ in range(gap):
                    container.insert(
                        i,
                        vec_type(
                            *[val_type(gap.fill_val) for _ in range(vec_length)],
                            **gap.kwargs,
                        ),
                    )

    @classmethod
    def convert_all_to_frames(
        cls, container: list, frame_type: type, vec_length: int
    ) -> None:
        """
        Given a container, replace each Gap with a frame of vectors of cells.
        `vec_length` sets the length of each resulting vector.

        Mutates inplace: `container`
        """
        vec_type = frame_type.elem_type
        val_type = vec_type.elem_type
        for i, elem in enumerate(container):
            if isinstance(elem, cls):
                container[i] = frame_type(
                    *[
                        vec_type(*[val_type(elem.fill_val) for _ in range(vec_length)])
                        for _ in range(elem)
                    ],
                    **elem.kwargs,
                )

    def __repr__(self):
        return f"{self.__class__.__name__}({int(self)})"


class HasId:
    """
    Has an id property which, when set, inserts a reference to
    self in `Globals.ids`.
    """

    @property
    def id(self):
        if not hasattr(self, "_id"):
            self._id = None
        return self._id

    @id.setter
    def id(self, new):
        self.set_id(new)

    def set_id(self, new):
        if new is not None:
            if not isinstance(new, str):
                raise ValueError(f"Invalid id, `{new}`. Ids must be strings.")
            Globals.ids[new] = self
            if new.startswith("G::"):
                Globals.global_ids[new] = self
        self._id = new
        return self


class HasHeader:
    """
    Has an header property which, when set, inserts a reference to
    self in `Globals.headers`.
    """

    @property
    def header(self):
        if not hasattr(self, "_header"):
            self._header = None
        return self._header

    @header.setter
    def header(self, new):
        self.set_header(new)

    def set_header(self, new):
        if new is not None:
            if not isinstance(new, str):
                raise ValueError(f"Invalid header, `{new}`. Headers must be strings.")
            Globals.headers[new] = self
            if new.startswith("G::"):
                Globals.global_headers[new] = self
        self._header = new
        return self


class HasMargin:
    empty = [None, None, None, None]

    def init_margin(self, margin, top, right, bottom, left) -> None:

        if (margin, top, right, bottom, left) == (None, None, None, None, None):
            self.margin_top = None
            self.margin_right = None
            self.margin_bottom = None
            self.margin_left = None
            return

        if not isinstance(margin, list):
            margin = [margin]
        if len(margin) == 1:
            margin *= 4
        elif len(margin) == 2:
            margin *= 2
        elif len(margin) == 3:
            margin += [None]

        self.margin_top = margin[0]
        self.margin_right = margin[1]
        self.margin_bottom = margin[2]
        self.margin_left = margin[3]

        if top is not None:
            self.margin_top = top
        if right is not None:
            self.margin_right = right
        if bottom is not None:
            self.margin_bottom = bottom
        if left is not None:
            self.margin_left = left

    @property
    def margin(self) -> list[int | None]:
        return [
            self.margin_top,
            self.margin_right,
            self.margin_bottom,
            self.margin_left,
        ]


class HasPadding:
    empty = [None, None, None, None]

    def init_padding(self, padding, top, right, bottom, left) -> None:

        if (padding, top, right, bottom, left) == (None, None, None, None, None):
            self.padding_top = None
            self.padding_right = None
            self.padding_bottom = None
            self.padding_left = None
            return

        if not isinstance(padding, list):
            padding = [padding]
        if len(padding) == 1:
            padding *= 4
        elif len(padding) == 2:
            padding *= 2
        elif len(padding) == 3:
            padding += [None]

        self.padding_top = padding[0]
        self.padding_right = padding[1]
        self.padding_bottom = padding[2]
        self.padding_left = padding[3]

        if top is not None:
            self.padding_top = top
        if right is not None:
            self.padding_right = right
        if bottom is not None:
            self.padding_bottom = bottom
        if left is not None:
            self.padding_left = left

    @property
    def padding(self) -> list[int | None]:
        return [
            self.padding_top,
            self.padding_right,
            self.padding_bottom,
            self.padding_left,
        ]

class HasBorder:
    """
    Child class is responsible for making sure each instance
    has variable, 'border_x' for each side.
    """

    empty = [None, None, None, None]
    negated = [False, False, False, False]
    default_weight = "thin"
    default_color = "000000"
    default = ("thin", "000000")
    valid_weights = (
        "dashDot",
        "dashDotDot",
        "dashed",
        "dotted",
        "double",
        "hair",
        "medium",
        "thick",
        "thin",
        "mediumDashDot",
        "mediumDashDotDot",
        "mediumDashed",
        "slantDashDot",
    )

    def init_border(self, border, top, right, bottom, left) -> None:
        """
        Processes the full border and individual sides, where
        individual sides take priority only if they are not None
        """
        cls = self.__class__
        self.border = border
        if top is not None:
            self.border_top = cls.interpret_single_value(top)
        if right is not None:
            self.border_right = cls.interpret_single_value(right)
        if bottom is not None:
            self.border_bottom = cls.interpret_single_value(bottom)
        if left is not None:
            self.border_left = cls.interpret_single_value(left)

        _ = self.border

    @property
    def border(self) -> list:
        for side in ["border_top", "border_right", "border_bottom", "border_left"]:
            if not hasattr(self, side):
                setattr(self, side, None)

        cls = self.__class__

        self.border_top = cls.interpret_single_value(self.border_top)
        self.border_right = cls.interpret_single_value(self.border_right)
        self.border_bottom = cls.interpret_single_value(self.border_bottom)
        self.border_left = cls.interpret_single_value(self.border_left)

        return [
            self.border_top,
            self.border_right,
            self.border_bottom,
            self.border_left,
        ]

    @border.setter
    def border(self, new: list) -> None:
        top, right, bottom, left = self.__class__.parse_arg(new)
        self.border_top = top
        self.border_right = right
        self.border_bottom = bottom
        self.border_left = left

    @classmethod
    def is_valid(cls, value: Any) -> bool:
        if value is None or value is False:
            return True
        if isinstance(value, tuple):
            if len(value) == 2:
                if isinstance(value[0], str) and isinstance(value[1], str):
                    if not value[1].startswith("#"):
                        return True
        return False

    @classmethod
    def interpret_single_value(cls, value: Any) -> tuple[Any, Any]:
        """
        Given a value intended to represent a single border side, interpret
        it to one of the following valid formats:
            - None   - unset, can be overriden
            - False  - override parent and remove border
            - ('<weight>' | None | False, '<hex color>' | None | False)

        Valid inputs for ``value``:
            None
            True    - converts to ``cls.default``
            False
            '<weight>'   - we can tell if the string is in list of valid weights
            '<hex color>'
            ('<weight>' | None | True | False,)
            ('<weight>' | None | True | False, '<hex color>' | None | True | False)
        """
        if cls.is_valid(value):
            return value

        # Treat 1-element tuple as single value
        if isinstance(value, tuple):
            if len(value) == 1:
                value = value[0]

        if value is True:
            return cls.default

        # If string, we can definitively conclude whether they were
        # referring to the weight or to the color.
        if isinstance(value, str):
            if value in cls.valid_weights:
                return (value, cls.default_color)
            else:
                return (cls.default_weight, value)

        # Now it must be a 2-element tuple
        if not isinstance(value, tuple):
            raise ValueError(f"Invalid border value, {value}")
        if not len(value) == 2:
            raise ValueError(f"Invalid border value, {value}")

        if value[0] is True:
            value = (cls.default_weight, value[1])

        if value[1] is True:
            value = (value[0], cls.default_color)

        for val in value:
            if not isinstance(val, str) and not val is None and not val is False:
                raise ValueError(f"Border weight/color values must be strings. {value} is invalid")

        if not value[0] in cls.valid_weights and not value[0] is None and not value[0] is False:
            raise ValueError(f"'{value[0]}' is not a valid weight")
        if not isinstance(value[1], str) and not value[1] is None and not value[1] is False:
            raise ValueError(f"'{value[1]}' is not a valid hex color")

        if isinstance(value[1], str):
            value = (value[0], value[1].lstrip("#"))
            if not len(value[1]) == 6:
                raise ValueError(f"Color value must be 6-character hex code. {value[1]} is invalid")

        return value

    @classmethod
    def parse_arg(
        cls, border: bool | Iterable | None
    ) -> list:
        """
        Designed to mimic CSS border logic. Returns a 4-element list
        describing the border of 4 sides, in the order: top, right, bottom,
        left. Elements can either be None, False, or a string representing weight.

        Arguments of True will default to 'thin' border

        Example arguments to outputs:
            True or 'thin':
                [ 'thin', 'thin', 'thin', 'thin' ]
            [ 'thin', False ]:
                [ 'thin', False, 'thin', False ]
            [ 'thick', 'thick', 'thick' ]:
                [ 'thick', 'thick', 'thick', False ]
        """
        if border is None or border == cls.empty:
            return cls.empty
        if border == cls.negated:
            return cls.negated

        if isinstance(border, tuple):
            if len(border) > 2:
                print(border)
                raise TypeError("Border must be a list")

        if not isinstance(border, list):
            border = [border]

        for i, elem in enumerate(border):
            border[i] = cls.interpret_single_value(elem)

        if len(border) == 1:
            border = border * 4
        elif len(border) == 2:
            border += border
        elif len(border) == 3:
            border += [False]

        assert len(border) == 4, "Border must be 4 elements. If you're reading this, an excelbird developer made a mistake"
        return list(border)

    def apply_border(self) -> None:
        if not hasattr(self, "__len__"):
            return
        if len(self) == 0 or self.border == [None, None, None, None]:
            return

        first = self[0]

        if len(self) == 1:
            if getattr(first, "is_empty", None) is True and hasattr(first, 'value'):
                first.value = ""
            first.border = self.border

        elif len(self) >= 2:
            mask = self.border_mask(*self.border)
            last = self[-1]
            vecs_in_between = self[1:-1]

            if getattr(first, "is_empty", None) is True and hasattr(first, 'value'):
                first.value = ""
            if getattr(last, "is_empty", None) is True and hasattr(last, 'value'):
                last.value = ""

            first.border = mask.first
            last.border = mask.last
            if len(self) > 2:
                for vec in vecs_in_between:
                    if getattr(vec, 'is_empty', None) is True and hasattr(vec, 'value'):
                        vec.value = ""
                    vec.border = mask.middle


class DotDict(dict):
    def __getattr__(self, key):
        try:
            return self[key]
        except KeyError:
            return super().__getattr__(key)


class Style(DotDict):
    def help(self) -> None:
        print(*self.keys(), sep="\n")


class ListIndexableById(list):
    """
    A simple child class of list that can accept an `id` string as a
    key to access elements.

    Each element MUST have an `id` property itself, before trying to
    access elements.
    """

    def key_to_idx(self, key) -> int:
        if isinstance(key, int):
            return key

        ids = [i.id if hasattr(i, "_id") else None for i in self]
        if key in ids:
            return ids.index(key)
        else:
            headers = [i.header if hasattr(i, "_header") else None for i in self]
            if key in headers:
                return headers.index(key)
            else:
                raise KeyError(f"Invalid key, {key}")

    def insert(self, index, new) -> None:
        index = self.key_to_idx(index)
        super().insert(index, new)

    def __setitem__(self, key, val) -> None:
        index = self.key_to_idx(key)
        super().__setitem__(index, val)

    def __getitem__(self, key) -> Any:
        if not isinstance(key, slice):
            key = self.key_to_idx(key)
        return super().__getitem__(key)

    def get(self, key, default=None) -> Any:
        try:
            return self[key]
        except Exception:
            return default

    def __repr__(self):
        # This shouldnt be here but I'm lazy
        return f"{self.__class__.__name__}({super().__repr__()})"


TLoc = TypeVar("TLoc", bound="Loc")


class Loc:
    """
    Stores y-x coordinates (ZERO-BASED) of a cell in Excel.
    Takes arguments in the order, `y`, `x`.
    ---
    Can be constructed from:
        String name of a cell in Excel: "B5"
        Another Loc
        2-element iterable of ints (y, x)
        Keywords, 'y' and 'x'
        2 positionals (y and x)

    Conversion to string (`str(my_loc)`) will return an Excel cell location "C7"
    """

    @overload
    def __init__(self, loc: TLoc, ws: None = None) -> None:
        ...

    @overload
    def __init__(self, loc: str | tuple[int, int] | list[int], ws: Worksheet) -> None:
        ...

    def __init__(
        self, loc: TLoc | str | tuple[int, int] | list[int], ws: Worksheet | None = None
    ) -> None:
        if isinstance(loc, Loc):
            self.y = loc.y
            self.x = loc.x
            self.ws = loc.ws

        elif isinstance(loc, (tuple, list)):
            self.y = loc[0]
            self.x = loc[1]
            self.ws = ws

        elif isinstance(loc, str):
            col_str, row_num = coordinate_from_string(loc)
            col_num = column_index_from_string(col_str)
            self.y = row_num - 1
            self.x = col_num - 1
            self.ws = ws
        else:
            raise ValueError(f"Invalid argument, {loc}")

        if self.ws is None:
            raise ValueError("A Loc must have a worksheet")

    @property
    def cell(self) -> str:
        return self.ws[self.cell_str]

    @property
    def col_letter(self) -> str:
        return get_column_letter(self.x + 1)

    @property
    def cell_str(self) -> str:
        return f"{get_column_letter(self.x+1)}{self.y+1}"

    @property
    def title_str(self) -> str:
        chars_to_trigger_quotes = [" ", "-"]
        if any(c in self.ws.title for c in chars_to_trigger_quotes):
            return "'" + self.ws.title + "'" + "!"
        return self.ws.title + "!"

    @property
    def full_str(self) -> str:
        return self.title_str + self.cell_str

    @property
    def column_dimensions(self):
        return self.ws.column_dimensions[self.col_letter]

    @property
    def row_dimensions(self):
        return self.ws.row_dimensions[self.y + 1]


class HasHelp:
    @classmethod
    def help(cls, doc: bool = False):
        doc_str = cls.__doc__
        help_str = cls.__help__
        help_nb_str = cls.__help_notebook__

        if doc_str is None and help_str is None and help_nb_str is None:
            res = f"Sorry, **`{cls.__name__}`** doesn't have any documentation yet."
        elif doc is True or help_str is None:
            res = doc_str
        elif help_str is not None:
            res = help_str

        from excelbird.util import is_notebook

        if is_notebook():
            from IPython.display import display, Markdown as md

            if help_nb_str is not None:
                display(md(help_nb_str))
            else:
                display(md(res))
        else:
            print(res)

    # Your class's help string
    __help__ = None
    # Custom help string for notebooks
    __help_notebook__ = None
