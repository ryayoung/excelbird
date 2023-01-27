from typing import overload, TypeVar
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter

TLoc = TypeVar("TLoc", bound="Loc")

class Loc:
    """
    Stores y-x coordinates (ZERO-BASED) of a cell in Excel.
    Takes arguments in the order, `y`, `x`.
    ---
    Can be constructed from:
        String name of a cell in Excel: "B5"
        Another Loc
        2-element iterable of ints (y, x)
        Keywords, 'y' and 'x'
        2 positionals (y and x)

    Conversion to string (`str(my_loc)`) will return an Excel cell location "C7"
    """

    @overload
    def __init__(self, loc: TLoc, ws: None = None) -> None:
        ...

    @overload
    def __init__(self, loc: str | tuple[int, int] | list[int], ws: Worksheet) -> None:
        ...

    def __init__(
        self, loc: TLoc | str | tuple[int, int] | list[int], ws: Worksheet | None = None
    ) -> None:
        if isinstance(loc, Loc):
            self.y = loc.y
            self.x = loc.x
            self.ws = loc.ws

        elif isinstance(loc, (tuple, list)):
            self.y = loc[0]
            self.x = loc[1]
            self.ws = ws

        elif isinstance(loc, str):
            col_str, row_num = coordinate_from_string(loc)
            col_num = column_index_from_string(col_str)
            self.y = row_num - 1
            self.x = col_num - 1
            self.ws = ws
        else:
            raise ValueError(f"Invalid argument, {loc}")

        if self.ws is None:
            raise ValueError("A Loc must have a worksheet")

    @property
    def cell(self) -> str:
        return self.ws[self.cell_str]

    @property
    def col_letter(self) -> str:
        return get_column_letter(self.x + 1)

    @property
    def cell_str(self) -> str:
        return f"{get_column_letter(self.x+1)}{self.y+1}"

    @property
    def title_str(self) -> str:
        chars_to_trigger_quotes = [" ", "-"]
        if any(c in self.ws.title for c in chars_to_trigger_quotes):
            return "'" + self.ws.title + "'" + "!"
        return self.ws.title + "!"

    @property
    def full_str(self) -> str:
        return self.title_str + self.cell_str

    @property
    def column_dimensions(self):
        return self.ws.column_dimensions[self.col_letter]

    @property
    def row_dimensions(self):
        return self.ws.row_dimensions[self.y + 1]
